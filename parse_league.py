"""
Pinball League Data Parser
--------------------------
Run this script at the end of each league season to convert your Excel
file into the JSON format used by the Pinball Stats web app.

USAGE:
  1. Edit the CONFIG section below (league name and file path)
  2. Open a terminal in this folder
  3. Run:  python parse_league.py
  4. The output file will appear in the data/ folder
  5. Push everything to GitHub:
       git add .
       git commit -m "Add Spring 2026 league data"
       git push
"""

import json
import os
import openpyxl

# ============================================================
# CONFIG — Edit these two lines before running
# ============================================================
LEAGUE_NAME = "Spring 2026"          # Display name shown in the app
EXCEL_FILE  = "League_RawData.xlsx"  # Filename of your Excel file
#                                      (must be in the same folder as this script)
# ============================================================


def parse_sheet(ws, week_num):
    """
    Parse one sheet (one week of play) from the workbook.
    Returns a list of dicts: {last, first, week, machine, score}
    """
    rows = list(ws.iter_rows(values_only=True))

    # --- Find the header row ---
    # It's the row where column 0 or 1 contains a player label like
    # "Player", "Last Name", or "Last"
    header_row_idx = None
    for i, row in enumerate(rows):
        if row[0] in ('Player', 'Last Name', 'Last') or \
           row[1] in ('Player', 'First Name', 'First'):
            header_row_idx = i
            break

    if header_row_idx is None:
        print(f"  WARNING: Could not find header row in Week {week_num}. Skipping.")
        return []

    # --- Extract machine names from the header row ---
    # Machine names appear every 3 columns starting at column index 2.
    # The last column is the weekly rank score total — we skip it.
    game_name_row = rows[header_row_idx]
    machines = []
    col = 2
    while col < len(game_name_row) - 1:
        name = game_name_row[col]
        if name and name not in ('Rank', 'Score', 'Rank Score'):
            machines.append((col, str(name).strip()))
        col += 3

    # --- Data rows start 4 rows after the header ---
    # Header row: "Player | Player | MachineName | ..."
    # +1 row:     "Last Name | First Name | ..."
    # +2 row:     "Score | Rank | Rank | ..."
    # +3 row:     "| | Points | ..."
    # +4 row:     First actual player data
    data_start = header_row_idx + 4

    records = []
    for row in rows[data_start:]:
        last  = row[0]
        first = row[1]

        # Skip blank rows or rows without a player name
        if not last or not first:
            continue

        # Skip players who didn't play (no scores at all — e.g. absent members)
        has_score = any(
            col_idx < len(row) and isinstance(row[col_idx], (int, float))
            for col_idx, _ in machines
        )
        if not has_score:
            continue

        # Extract one record per machine played
        for col_idx, machine in machines:
            score = row[col_idx] if col_idx < len(row) else None
            if score and isinstance(score, (int, float)):
                records.append({
                    "last":    str(last).strip(),
                    "first":   str(first).strip(),
                    "week":    week_num,
                    "machine": machine,
                    "score":   int(score)
                })

    return records


def main():
    # --- Locate the Excel file ---
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, EXCEL_FILE)

    if not os.path.exists(excel_path):
        print(f"ERROR: Cannot find '{EXCEL_FILE}' in {script_dir}")
        print("Make sure the Excel file is in the same folder as this script.")
        return

    print(f"Reading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")

    # --- Parse every sheet ---
    all_records = []
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        week_num = i + 1
        recs = parse_sheet(ws, week_num)
        print(f"  Week {week_num} ({sheet_name}): {len(recs)} score records")
        all_records.extend(recs)

    if not all_records:
        print("ERROR: No records were parsed. Check your Excel file.")
        return

    # --- Build unique player list (for the app's autocomplete) ---
    seen = set()
    players = []
    for r in all_records:
        key = (r["last"], r["first"])
        if key not in seen:
            seen.add(key)
            players.append({"last": r["last"], "first": r["first"]})
    players.sort(key=lambda p: (p["last"], p["first"]))

    # --- Assemble final JSON output ---
    output = {
        "league":  LEAGUE_NAME,
        "weeks":   len(wb.sheetnames),
        "players": players,
        "records": all_records
    }

    # --- Save to data/ folder ---
    # File name is auto-generated from league name: "Spring 2026" -> "spring2026.json"
    file_name = LEAGUE_NAME.lower().replace(" ", "") + ".json"
    output_dir = os.path.join(script_dir, "data")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, file_name)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2)

    print(f"\nSuccess! {len(all_records)} records written to: data/{file_name}")
    print(f"Players found: {len(players)}")
    print(f"\nNext steps:")
    print(f"  git add .")
    print(f"  git commit -m \"Add {LEAGUE_NAME} league data\"")
    print(f"  git push")


if __name__ == "__main__":
    main()
